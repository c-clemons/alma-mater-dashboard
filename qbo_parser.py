"""
QBO Parser Module
Reads QuickBooks Online export files (.xlsx) and extracts P&L and Balance Sheet data
into a standardized format for use in the Streamlit app.
"""

import openpyxl
from typing import Dict, Tuple, Optional


# QBO P&L: search patterns -> standardized label
# The parser searches column A for these patterns to find correct rows dynamically.
QBO_PL_SEARCH = [
    (["total for 401000", "total 401000 dtc", "total dtc"], "DTC Revenue"),
    (["total for 402000", "total 402000 wholesale", "total wholesale"], "Wholesale Revenue"),
    (["total for income", "total income"], "Total Revenue"),
    (["total for cost of goods sold", "total cost of goods sold"], "Total COGS"),
    (["gross profit"], "Gross Profit"),
    (["total for 601000", "total 601000 sales"], "Sales & Marketing"),
    (["total for 602000 payroll", "total for 603000 payroll", "602000 payroll", "603000 payroll"], "Payroll"),
    (["603000 software", "602000 software"], "Software"),
    (["total for 604000", "total 604000 professional"], "Professional Fees"),
    (["605000 travel", "606000 travel"], "Travel"),
    (["606000 meals", "605000 meals"], "Meals"),
    (["607000 entertainment"], "Entertainment"),
    (["608000 insurance", "609000 insurance"], "Insurance"),
    (["609000 office furniture", "611000 office furniture"], "Office Furniture & Equipment"),
    (["610000 office supplies", "612000 office supplies"], "Office Supplies"),
    (["613000 bank", "608000 bank"], "Bank Fees"),
    (["total for 614000", "614000 rent"], "Rent & Lease"),
    (["617000 utilities", "615000 utilities"], "Utilities"),
    (["618000 repairs", "616000 repairs"], "Repairs & Maintenance"),
    (["paypal fees", "613000 job supplies", "job supplies"], "PayPal Fees / Other"),
    (["total for expenses", "total expenses"], "Total Expenses"),
    (["net operating income"], "Net Operating Income"),
    (["810000 depreciation", "depreciation exp", "depreciation"], "Depreciation"),
    (["900000 other", "other income", "total for other"], "Other Income/(Expense)"),
    (["net income"], "Net Income"),
]

# All standardized labels (preserves iteration order for build_actuals_dataframe)
QBO_PL_LABELS = [label for _, label in QBO_PL_SEARCH]

# QBO BS search for cash row
QBO_BS_CASH_SEARCH = ["total bank accounts", "total for bank accounts"]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4,
    "june": 6, "july": 7, "august": 8, "september": 9,
    "october": 10, "november": 11, "december": 12,
}


def parse_qbo_headers(ws) -> Dict[Tuple[int, int], int]:
    """Parse QBO column headers to determine month mapping.
    Returns dict: {(year, month_num): column_index}
    """
    headers = {}
    for c in range(2, 50):
        val = ws.cell(row=5, column=c).value
        if val is None:
            continue
        val_str = str(val).strip()
        if val_str == "Total":
            break
        parts = val_str.split()
        if len(parts) == 2:
            mo_str = parts[0].lower()
            yr_str = parts[1]
            if mo_str in MONTH_MAP and yr_str.isdigit():
                headers[(int(yr_str), MONTH_MAP[mo_str])] = c
    return headers


def _find_row_by_search(ws, search_terms, max_row=200):
    """Find a row in column A matching any of the search terms (case-insensitive)."""
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        for term in search_terms:
            if term in val_lower:
                return r
    return None


def _build_pl_row_map(ws):
    """Dynamically build QBO P&L row mapping by searching column A."""
    row_map = {}  # {label: row_number}
    for search_terms, label in QBO_PL_SEARCH:
        row = _find_row_by_search(ws, search_terms)
        if row:
            row_map[label] = row
    return row_map


def parse_qbo_file(file_bytes) -> Dict:
    """
    Parse a QBO export file (uploaded via Streamlit file_uploader).

    Args:
        file_bytes: BytesIO object from st.file_uploader

    Returns:
        Dict with keys:
            'pl_data': {label: {(year, month): value}}
            'cash_data': {(year, month): cash_balance}
            'last_year': int
            'last_month': int
            'latest_cash': float
            'months_found': list of (year, month) tuples
    """
    wb = openpyxl.load_workbook(file_bytes, data_only=True)

    # Find sheet names (handle "Profit and Loss" or "PL", "Balance Sheet" or "BS")
    pl_name = None
    bs_name = None
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if nl in ("pl", "p&l", "p & l") or ("profit" in nl and "loss" in nl):
            pl_name = name
        if nl in ("bs",) or ("balance" in nl and "sheet" in nl):
            bs_name = name

    if not pl_name:
        raise ValueError(f"Could not find P&L sheet. Found: {wb.sheetnames}")

    pl_ws = wb[pl_name]

    # Parse P&L headers
    pl_headers = parse_qbo_headers(pl_ws)
    if not pl_headers:
        raise ValueError("Could not parse month headers from P&L sheet row 5")

    # Dynamically find P&L rows by searching account names in column A
    row_map = _build_pl_row_map(pl_ws)

    # Extract P&L data
    pl_data = {}
    for label, qbo_row in row_map.items():
        pl_data[label] = {}
        for (yr, mo), col in pl_headers.items():
            val = pl_ws.cell(row=qbo_row, column=col).value
            pl_data[label][(yr, mo)] = float(val) if val is not None else 0.0

    # Extract BS data if available
    cash_data = {}
    if bs_name:
        bs_ws = wb[bs_name]
        bs_headers = parse_qbo_headers(bs_ws)
        cash_row = _find_row_by_search(bs_ws, QBO_BS_CASH_SEARCH)
        if cash_row:
            for (yr, mo), col in bs_headers.items():
                val = bs_ws.cell(row=cash_row, column=col).value
                cash_data[(yr, mo)] = float(val) if val is not None else 0.0

    # Determine last actuals month
    last_year, last_month = max(pl_headers.keys())
    latest_cash = cash_data.get((last_year, last_month), 0.0)

    return {
        'pl_data': pl_data,
        'cash_data': cash_data,
        'last_year': last_year,
        'last_month': last_month,
        'latest_cash': latest_cash,
        'months_found': sorted(pl_headers.keys()),
    }


def build_actuals_dataframe(pl_data: Dict, year: int, max_month: int = 12) -> Dict[str, list]:
    """
    Build a monthly actuals dict for a given year from parsed QBO data.

    Returns dict with P&L line items as keys and lists of 12 monthly values.
    """
    result = {}
    for label in QBO_PL_LABELS:
        monthly = []
        for mo in range(1, 13):
            if mo <= max_month:
                val = pl_data.get(label, {}).get((year, mo), 0.0)
            else:
                val = 0.0
            monthly.append(val)
        result[label] = monthly
    return result


def actuals_to_pl_format(actuals: Dict[str, list]) -> list:
    """
    Convert QBO actuals dict to a list of monthly dicts matching the P&L format
    used by the dashboard (same keys as generate_monthly_pl output).
    """
    months = MONTHS
    data = []
    for m in range(12):
        dtc_rev = actuals.get("DTC Revenue", [0]*12)[m]
        ws_rev = actuals.get("Wholesale Revenue", [0]*12)[m]
        total_rev = actuals.get("Total Revenue", [0]*12)[m]
        total_cogs = actuals.get("Total COGS", [0]*12)[m]
        gross_profit = actuals.get("Gross Profit", [0]*12)[m]
        total_expenses = actuals.get("Total Expenses", [0]*12)[m]
        net_op_income = actuals.get("Net Operating Income", [0]*12)[m]

        # Map QBO "Total Expenses" to our OpEx (it includes S&M, payroll, etc.)
        data.append({
            'Month': months[m],
            'DTC Revenue': dtc_rev,
            'Wholesale Revenue': ws_rev,
            'Total Revenue': total_rev,
            'Total COGS': total_cogs,
            'Gross Profit': gross_profit,
            'Gross Margin %': (gross_profit / total_rev * 100) if total_rev > 0 else 0,
            'Total OpEx': total_expenses,
            'EBITDA': net_op_income,
            'EBITDA Margin %': (net_op_income / total_rev * 100) if total_rev > 0 else 0,
            # Detailed breakdown
            'Sales & Marketing': actuals.get("Sales & Marketing", [0]*12)[m],
            'Payroll': actuals.get("Payroll", [0]*12)[m],
            'Software': actuals.get("Software", [0]*12)[m],
            'Professional Fees': actuals.get("Professional Fees", [0]*12)[m],
            'Travel': actuals.get("Travel", [0]*12)[m],
            'Meals': actuals.get("Meals", [0]*12)[m],
            'Entertainment': actuals.get("Entertainment", [0]*12)[m],
            'Insurance': actuals.get("Insurance", [0]*12)[m],
            'Rent & Lease': actuals.get("Rent & Lease", [0]*12)[m],
            'Utilities': actuals.get("Utilities", [0]*12)[m],
            'Bank Fees': actuals.get("Bank Fees", [0]*12)[m],
            'Office Supplies': actuals.get("Office Supplies", [0]*12)[m],
            'Depreciation': actuals.get("Depreciation", [0]*12)[m],
            'Net Income': actuals.get("Net Income", [0]*12)[m],
        })
    return data


def serialize_qbo_data(parsed: Dict) -> Dict:
    """Convert parsed QBO data to JSON-serializable format for session state / persistence."""
    pl_serialized = {}
    for label, month_data in parsed['pl_data'].items():
        pl_serialized[label] = {f"{yr}_{mo}": val for (yr, mo), val in month_data.items()}

    cash_serialized = {f"{yr}_{mo}": val for (yr, mo), val in parsed['cash_data'].items()}

    return {
        'pl_data': pl_serialized,
        'cash_data': cash_serialized,
        'last_year': parsed['last_year'],
        'last_month': parsed['last_month'],
        'latest_cash': parsed['latest_cash'],
        'months_found': [f"{yr}_{mo}" for yr, mo in parsed['months_found']],
    }


def deserialize_qbo_data(data: Dict) -> Dict:
    """Convert JSON-serialized QBO data back to the parsed format."""
    if not data:
        return None

    pl_data = {}
    for label, month_data in data.get('pl_data', {}).items():
        pl_data[label] = {}
        for key, val in month_data.items():
            yr, mo = key.split('_')
            pl_data[label][(int(yr), int(mo))] = val

    cash_data = {}
    for key, val in data.get('cash_data', {}).items():
        yr, mo = key.split('_')
        cash_data[(int(yr), int(mo))] = val

    months_found = []
    for key in data.get('months_found', []):
        yr, mo = key.split('_')
        months_found.append((int(yr), int(mo)))

    return {
        'pl_data': pl_data,
        'cash_data': cash_data,
        'last_year': data.get('last_year', 2025),
        'last_month': data.get('last_month', 12),
        'latest_cash': data.get('latest_cash', 0),
        'months_found': months_found,
    }
